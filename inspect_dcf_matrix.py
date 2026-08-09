import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\omkar\OneDrive\Documents\Portfolio\NexGen_Tech_Financial_Model_and_DCF_Valuation.xlsx', data_only=True)

print("Sheet names:", wb.sheetnames)

for sheetname in wb.sheetnames:
    print(f"\n--- SHEET: {sheetname} ---")
    ws = wb[sheetname]
    for row in ws.iter_rows(values_only=True):
        row_str = [str(c) if c is not None else "" for c in row]
        if any(row_str):
            print(" | ".join(row_str[:15]))
